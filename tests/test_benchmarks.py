import datetime
import json
import os
import unittest
from pathlib import Path
from typing import Union

import numpy as np
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from osgeo import gdal

from eotimeseriesviewer import initAll, DIR_REPO
from eotimeseriesviewer.force import FORCEUtils
from eotimeseriesviewer.forceinputs import FindFORCEProductsTask
from eotimeseriesviewer.main import EOTimeSeriesViewer
from eotimeseriesviewer.processing.processingalgorithms import ReadTemporalProfiles
from eotimeseriesviewer.tests import EOTSVTestCase, start_app
from eotimeseriesviewer.timeseries.source import TimeSeriesSource
from eotimeseriesviewer.timeseries.tasks import hasValidPixel
from qgis import processing
from qgis.core import QgsApplication, QgsProcessingAlgRunnerTask, QgsProject, QgsRasterLayer, QgsTaskManager
from qgis.core import QgsCoordinateReferenceSystem, QgsRectangle

start_app()
initAll()

DIR_BENCHMARKS = DIR_REPO / 'benchmarks'
os.makedirs(DIR_BENCHMARKS, exist_ok=True)

FORCE_CUBE = Path(os.environ.get('FORCE_CUBE', '-'))
FORCE_MOSAICS = FORCE_CUBE / 'mosaic'


def get_workbook(path_xlsx: Union[str, Path]) -> Workbook:
    path_xlsx = Path(path_xlsx)
    if path_xlsx.is_file():
        book = load_workbook(filename=path_xlsx.as_posix())
    else:
        book = Workbook()
    for s in book.sheetnames[:]:
        del book[s]
    return book


def get_sheet(name: str, book: Workbook) -> Worksheet:
    if name in book.sheetnames:
        sheet = book[name]
        sheet.delete_cols(1, sheet.max_column)
    else:
        sheet = book.create_sheet(name)
    return sheet


@unittest.skipIf(EOTSVTestCase.runsInCI(), 'Benchmark Tests. Not to run in CI')
class BenchmarkTestCase(EOTSVTestCase):

    @unittest.skipIf(not FORCE_MOSAICS.is_dir(), 'Missing FORCE_CUBE')
    def test_load_mosaic_overlap(self):
        path_files = DIR_BENCHMARKS / 'benchmark_load_mosaics_files.json'
        path_json = DIR_BENCHMARKS / 'benchmark_load_mosaics_overlap.json'
        files = None

        import tqdm

        if not path_files.is_file():
            task = FindFORCEProductsTask('BOA', FORCE_MOSAICS, dateMax='1986-12-31')
            task.run_task_manager()
            files = [f.as_posix() for f in task.files()]
            with open(path_files, 'w') as f:
                json.dump(files, f)
        else:
            with open(path_files, 'r') as f:
                files = json.load(f)

        assert isinstance(files, list)
        files = files[0:min(100, len(files))]

        n = len(files)
        self.assertTrue(n > 0)
        duration_lyr = []
        duration_gdal = []
        sample_size = 256

        crs = QgsCoordinateReferenceSystem('EPSG:3035')
        extent_wkt = 'Polygon ((4553985.05148862022906542 3290907.41572995623573661, 4555977.23898862022906542 3290907.41572995623573661, 4555977.23898862022906542 3292899.60322995623573661, 4553985.05148862022906542 3292899.60322995623573661, 4553985.05148862022906542 3290907.41572995623573661))'
        extent = QgsRectangle.fromWkt(extent_wkt)

        def get_overlap(path, use_gdal: bool):
            t = datetime.datetime.now()
            success, err = hasValidPixel(path, crs, extent, use_gdal=use_gdal)
            dt = datetime.datetime.now() - t
            return success, err, dt.total_seconds()

        # 20210609_LEVEL2_LND08_BOA.vrt
        #
        with tqdm.tqdm(files) as pbar:
            for file in pbar:
                file = str(file)
                success1, err1, t1 = get_overlap(file, False)
                success2, err2, t2 = get_overlap(file, True)
                self.assertEqual(success1, success2)
                if success1:
                    duration_lyr.append(t1)
                    duration_gdal.append(t2)
                pbar.update(1)

        if path_json.is_file():
            with open(path_json, 'r') as f:
                RESULTS = json.load(f)
        else:
            RESULTS = dict()

        dur_gdal = np.asarray(duration_gdal)
        dur_ts = np.asarray(duration_lyr)

        descr = 'test mosaic overlap'
        INFO = {'description': descr,
                'n_files': n,
                'root': FORCE_MOSAICS.as_posix(),
                't_lyr_total': float(dur_ts.sum()),
                't_lyr_mean': float(dur_ts.mean()),
                't_lyr_min': float(dur_ts.min()),
                't_lyr_max': float(dur_ts.max()),
                't_lyr_std': float(dur_ts.std()),
                't_gdal_total': float(dur_gdal.sum()),
                't_gdal_mean': float(dur_gdal.mean()),
                't_gdal_min': float(dur_gdal.min()),
                't_gdal_max': float(dur_gdal.max()),
                't_gdal_std': float(dur_gdal.std())
                }

        RESULTS[descr] = INFO

        with open(path_json, 'w') as f:
            json.dump(RESULTS, f, indent=2)

    @unittest.skipIf(not FORCE_MOSAICS.is_dir(), 'Missing FORCE_CUBE')
    def test_load_mosaic_sources(self):
        path_files = DIR_BENCHMARKS / 'benchmark_load_mosaics_files.json'
        path_json = DIR_BENCHMARKS / 'benchmark_load_mosaics_results.json'
        files = None

        import tqdm

        if not path_files.is_file():
            task = FindFORCEProductsTask('BOA', FORCE_MOSAICS, dateMax='1986-12-31')
            task.run_task_manager()
            files = [f.as_posix() for f in task.files()]
            with open(path_files, 'w') as f:
                json.dump(files, f)
        else:
            with open(path_files, 'r') as f:
                files = json.load(f)

        assert isinstance(files, list)
        files = files[0:min(100, len(files))]

        n = len(files)
        self.assertTrue(n > 0)
        duration_ts = []
        duration_gdal = []

        def t1_get_ts_source(path):
            t = datetime.datetime.now()
            src = TimeSeriesSource.create(str(file))
            self.assertIsInstance(src, TimeSeriesSource)
            return datetime.datetime.now() - t

        def t2_get_gdal_with_md(path):
            t = datetime.datetime.now()
            ds = gdal.Open(str(path))
            self.assertIsInstance(ds, gdal.Dataset)
            mds = []
            for d in ds.GetMetadataDomainList():
                mds.append(ds.GetMetadata_Dict(d))
            for b in range(ds.RasterCount):
                band = ds.GetRasterBand(b + 1)
                for d in ds.GetMetadataDomainList():
                    mds.append(band.GetMetadata_Dict(d))
            return datetime.datetime.now() - t

        # 20210609_LEVEL2_LND08_BOA.vrt
        #
        with tqdm.tqdm(files) as pbar:
            for file in pbar:
                duration_ts.append(t1_get_ts_source(file))
                duration_gdal.append(t2_get_gdal_with_md(file))
                pbar.update(1)

        if path_json.is_file():
            with open(path_json, 'r') as f:
                RESULTS = json.load(f)
        else:
            RESULTS = dict()

        dur_gdal = np.asarray([d.total_seconds() for d in duration_gdal])
        dur_ts = np.asarray([d.total_seconds() for d in duration_ts])

        descr = 'load mosaics, improved TimeSeriesSource.create(file)'
        INFO = {'description': descr,
                'n_files': n,
                'root': FORCE_MOSAICS.as_posix(),
                't_ts_total': float(dur_ts.sum()),
                't_ts_mean': float(dur_ts.mean()),
                't_ts_min': float(dur_ts.min()),
                't_ts_max': float(dur_ts.max()),
                't_ts_std': float(dur_ts.std()),
                't_gdal_total': float(dur_gdal.sum()),
                't_gdal_mean': float(dur_gdal.mean()),
                't_gdal_min': float(dur_gdal.min()),
                't_gdal_max': float(dur_gdal.max()),
                't_gdal_std': float(dur_gdal.std())
                }

        RESULTS[descr] = INFO

        with open(path_json, 'w') as f:
            json.dump(RESULTS, f, indent=2)

    @unittest.skipIf(not FORCE_CUBE.is_dir(), 'Missing FORCE_CUBE')
    def test_benchmark_load_eotsv(self):

        path_results = self.createTestOutputDirectory() / 'benchmark_load_eotsv.json'

        if not path_results.is_file():

            tile_ids = [self.get_example_tiledir()]

            task = FindFORCEProductsTask('BOA', FORCE_CUBE, tile_ids=tile_ids)
            task.run_task_manager()

            files = task.files()

            n_threads = [2, 4, 6]
            n_files = [50, 100, 200, 400, 800]
            n_files = [n for n in n_files if n < len(files)]
            results = {'force_cube': str(FORCE_CUBE),
                       'tile_ids': tile_ids,
                       'results': []
                       }
            for nt in n_threads:
                for nf in n_files:
                    files_to_load = files[0:nf]
                    eotsv = EOTimeSeriesViewer()
                    eotsv.ui.show()
                    self.taskManagerProcessEvents()
                    t0 = datetime.datetime.now()
                    eotsv.addTimeSeriesImages(files_to_load)
                    self.taskManagerProcessEvents()
                    dt = datetime.datetime.now() - t0
                    d_result = {'n_threads': nt,
                                'n_files': nf,
                                'seconds': dt.total_seconds()}

                    results['results'].append(d_result)
                    print(d_result)
                    eotsv.close()
                    with open(path_results, 'w') as f:
                        json.dump(results, f, indent=2)
        else:
            with open(path_results, 'r') as f:
                results = json.load(f)
            print(f'Benchmark results: {path_results}')
            for (nt, nf), dt in results.items():
                print(f'threads={nt}, files={nf}: {dt}')

    def get_example_tiledir(self):

        tiles = [d.name for d in FORCEUtils.tileDirs(FORCE_CUBE)]
        assert len(tiles) > 0
        if 'X0066_Y0058' in tiles:
            tile_id = 'X0066_Y0058'
        else:
            tile_id = tiles[0]
        return tile_id

    @unittest.skipIf(not FORCE_CUBE.is_dir(), 'Missing FORCE_CUBE')
    def test_load_force_processingalgorithm_benchmark(self):

        tile_id = self.get_example_tiledir()

        thin_bottom_border = Border(bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        top_alignment = Alignment(vertical="top")

        task = FindFORCEProductsTask('BOA', FORCE_CUBE, tile_ids=[tile_id])
        task.run_task_manager()
        all_files = task.files()

        assert len(all_files) > 0

        dir_outputs = self.createTestOutputDirectory()

        project = QgsProject()

        path_results = dir_outputs / 'benchmark.json'

        if not path_results.is_file():
            n_threads = [2, 4, 6, 8]
            n_files = [1, 25, 50, 100, 150, 200, 250, 300]
            n_files = [n for n in n_files if n < len(all_files)]
            n_points = [1, 5, 10, 50]

            results = []

            for npts in n_points:
                path_random_points = dir_outputs / f'random_points_{npts}.geojson'
                path_random_points.unlink(missing_ok=True)

                for f in all_files:
                    lyr = QgsRasterLayer(f.as_posix())
                    project.addMapLayer(lyr)

                    context, feedback = self.createProcessingContextFeedback()
                    context.setProject(project)

                    assert lyr.isValid()
                    processing.run("native:randompointsinextent", {
                        'EXTENT': lyr.extent(),
                        'POINTS_NUMBER': npts,
                        'MIN_DISTANCE': 30,
                        'TARGET_CRS': lyr.crs(),
                        'MAX_ATTEMPTS': 200,
                        'OUTPUT': str(path_random_points)},
                                   context=context, feedback=feedback)

                    break

                for nf in n_files:
                    for nt in n_threads:
                        path_output = dir_outputs / f'files_{nf}_threads_{nt}.geojson'
                        path_output.unlink(missing_ok=True)

                        alg = ReadTemporalProfiles()
                        conf = {}
                        alg.initAlgorithm(conf)
                        files = [str(f) for f in all_files[0: nf]]
                        param = {
                            alg.INPUT: path_random_points.as_posix(),
                            alg.TIMESERIES: files,
                            alg.N_THREADS: nt,
                            alg.FIELD_NAME: 'tp',
                            alg.OUTPUT: path_output.as_posix(),
                        }

                        context, feedback = self.createProcessingContextFeedback()
                        context.setProject(project)

                        task = QgsProcessingAlgRunnerTask(alg, param, context, feedback)
                        self.assertTrue(task.canCancel())
                        tm: QgsTaskManager = QgsApplication.taskManager()

                        t0 = datetime.datetime.now()
                        tm.addTask(task)
                        self.taskManagerProcessEvents()
                        dt = datetime.datetime.now() - t0

                        d_result = {'n_files': nf,
                                    'n_threads': nt,
                                    'n_points': npts,
                                    'seconds': dt.total_seconds()}

                        results.append(d_result)
                        print(d_result)
                        with open(path_results, 'w') as f:
                            json.dump(results, f, indent=4)
        else:
            with open(path_results, 'r') as f:
                results = json.load(f)

            path_xlsx = path_results.parent / (path_results.stem + '.xlsx')

            wb = get_workbook(path_xlsx)
            sheet = get_sheet('benchmark', wb)
            sheet.cell(1, 1, f'Benchmark {path_results}')
            sheet.cell(2, 1)
            matrix = dict()
            for e in results:
                nf = e['n_files']
                nt = e['n_threads']
                np = e['n_points']
                k = (nf, nt, np)
                matrix[k] = e['seconds']

            n_files = sorted(set([k[0] for k in matrix.keys()]))
            n_threads = sorted(set([k[1] for k in matrix.keys()]))
            n_points = sorted(set([k[2] for k in matrix.keys()]))

            wb.save(path_xlsx.as_posix())


if __name__ == '__main__':
    unittest.main()
